#!/usr/bin/env python3
"""Local Excel application tracker - replaces Notion entirely.

Notion's API had a persistent, unresolved format quirk (parent.data_source_id
vs page_id) that cost a few wasted tool calls almost every time a job got
logged, plus a live network dependency for something that's fundamentally
just "is this company in my list, and if not, add a row." A local .xlsx
file removes both problems, and it's directly openable/reviewable by the
user without a browser - which is the whole reason for the switch.

This is also now the single source of truth for "already engaged with this
company" - applied_companies.json's old fast-mirror role is folded in here
instead of keeping two separate dedup files in sync.

The Excel sheet is the tracker, not a document store - the actual resume
and job description PDFs live as plain files on disk (a spreadsheet is a
poor place to browse/open a PDF from), and the sheet just holds a
clickable link to each. Every compiled resume also gets a copy dropped in
resumes/by_company/<Company>_resume_<ID>.pdf, and the job description
(plain text on disk) gets rendered to resumes/by_company/<Company>_<ID>.pdf
via text_to_pdf.py - one flat, human-browsable folder for both, since the
per-job resumes/<job_id>/ layout is convenient for the pipeline but not
something a person wants to hunt through in Finder. <ID> is one persistent
5-digit number per job (see get_or_create_file_id(), stored on the job's
own jobs.json record) shared by BOTH files, so a resume and its JD are
recognizably a pair at a glance - company name alone isn't unique (the
same company can generate more than one job row over time), which is why
a bare <Company>_resume.pdf would silently collide with an earlier job's.

Usage:
  python3 tracker.py list-companies [--out PATH]
    Writes a JSON array of every tracked company name (normalized) to
    --out (default: stdout). Used as the --skip-companies input for
    write_discovered_jobs.py - this whole step needs no agent/LLM call.

  python3 tracker.py check --company "Acme Corp"
    Prints "true" or "false" and exits 0/1 accordingly.

  python3 tracker.py add --job-id JOB_ID --company "Acme Corp" --role "ML Engineer" \
      --status "Ready for review" [--location L] [--source S] \
      [--url U] [--resume-path P] [--jd-path P] [--date-posted D] \
      [--work-type remote|hybrid|onsite] [--salary S] [--notes N]
    Appends one row. Always pass --job-id (the jobs.json id) for a real
    job - it's what lets the resume and JD PDFs share one persistent
    5-digit number. --resume-path is the compiled resumes/<job_id>/
    resume.pdf - it gets copied to resumes/by_company/<Company>_resume_<ID>.pdf
    and that copy is what the sheet links to. --jd-path is the job's
    resumes/<job_id>/jd_full.txt (plain text) - it gets rendered to
    resumes/by_company/<Company>_<ID>.pdf and that PDF is what the
    sheet links to. Creates the workbook with headers if it doesn't exist yet.

    Note: this only ever logs up to "ready for review" - the agent never
    clicks Submit, so it can't know if/when a real submission happened.
    The Status column's automated entries (Ready for review,
    Blocked-CAPTCHA, etc.) are a starting point - update it by hand as you
    actually submit and hear back (Submitted / Interviewing / Rejected /
    Offer), same as any manually-kept job tracker.
"""
import argparse
import fcntl
import json
import random
import re
import shutil
import sys
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

import text_to_pdf
from jobs_lock import locked_jobs_for_write

ROOT = Path(__file__).parent.parent
TRACKER_FILE = ROOT / "application_tracker.xlsx"
# Same reasoning/pattern as scripts/jobs_lock.py: cmd_add and
# cmd_update_status each do a plain load-workbook / mutate / save-workbook
# round trip. Two tracker.py processes doing that at the same moment (now
# a real possibility - jobs run concurrently, and each independently logs
# to the tracker when it reaches ready-for-review) is a read-modify-write
# race - whichever saves last wins and silently discards the other's row.
# Observed live: two Capital One rows with genuinely different titles both
# ended up "Submitted" in the sheet despite the user only actually
# submitting one - a symptom of a *different* bug (mark_submitted missing
# --role), but it demonstrates concurrent tracker writes are a real,
# already-occurring pattern, not a hypothetical one.
TRACKER_LOCK_FILE = TRACKER_FILE.with_suffix(".xlsx.lock")
BY_COMPANY_DIR = ROOT / "resumes" / "by_company"
SHEET_NAME = "Applications"
COLUMNS = ["Company", "Role", "Status", "Job Posting Date", "Date Applied",
           "Location", "Address Used", "Work Type", "Salary", "Source",
           "Job Link", "Resume", "Job Description", "Notes"]


@contextmanager
def locked_tracker_for_write():
    """Exclusive lock spanning a full load-workbook/mutate/save-workbook
    round trip. Blocks until any other reader or writer is done."""
    TRACKER_LOCK_FILE.touch(exist_ok=True)
    with open(TRACKER_LOCK_FILE, "r+") as lockfile:
        fcntl.flock(lockfile, fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(lockfile, fcntl.LOCK_UN)


@contextmanager
def locked_tracker_for_read():
    """Shared lock for a read-only pass - multiple readers can hold this
    at once, but it still blocks until any in-progress writer is done, so
    a reader can never observe a half-written .xlsx (openpyxl's save()
    isn't atomic - it streams a new zip archive directly to the target
    path, not via a temp-file-then-rename)."""
    TRACKER_LOCK_FILE.touch(exist_ok=True)
    with open(TRACKER_LOCK_FILE, "r+") as lockfile:
        fcntl.flock(lockfile, fcntl.LOCK_SH)
        try:
            yield
        finally:
            fcntl.flock(lockfile, fcntl.LOCK_UN)


def normalize_company(name) -> str:
    name = str(name or "").lower()
    name = re.sub(r"\b(inc|llc|corp|corporation|ltd|co|company|group|technologies|technology)\b\.?", "", name)
    name = re.sub(r"[^a-z0-9]+", "", name)
    return name


def sanitize_filename(name) -> str:
    name = re.sub(r'[\\/:*?"<>|]', "", str(name or "")).strip()
    return name or "company"


def ensure_workbook():
    if TRACKER_FILE.exists():
        return load_workbook(TRACKER_FILE)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(COLUMNS)
    for col_idx, header in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(header) + 2)
    return wb


def all_rows(wb) -> list[dict]:
    ws = wb[SHEET_NAME]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(COLUMNS, row)))
    return rows


def cmd_list_companies(args) -> None:
    with locked_tracker_for_read():
        wb = ensure_workbook()
        companies = sorted({normalize_company(r["Company"]) for r in all_rows(wb) if r.get("Company")})
    output = json.dumps(companies, indent=2)
    if args.out:
        Path(args.out).write_text(output)
        print(f"wrote {len(companies)} tracked companies -> {args.out}")
    else:
        print(output)


def cmd_check(args) -> None:
    with locked_tracker_for_read():
        wb = ensure_workbook()
        tracked = {normalize_company(r["Company"]) for r in all_rows(wb) if r.get("Company")}
    found = normalize_company(args.company) in tracked
    print("true" if found else "false")
    sys.exit(0 if found else 1)


FILE_ID_DIGITS = 5


def get_or_create_file_id(job_id: str) -> str:
    """Every job gets one persistent 5-digit ID, generated once and
    reused for BOTH its resume and job-description filenames in
    resumes/by_company/ - <Company>_resume_<ID>.pdf and <Company>_<ID>.pdf
    share the same number, so the two files are recognizably a pair at a
    glance (previously each file got its own independent random suffix,
    with no way to tell a resume and its JD belonged to the same job).
    Stored on the job's own jobs.json record (file_id field) so a retry
    or a later `tracker.py add` re-run reuses the same number instead of
    minting a new one."""
    with locked_jobs_for_write() as data:
        job = next((j for j in data["jobs"] if j["id"] == job_id), None)
        if job is None:
            raise SystemExit(f"no job found with id {job_id!r} (--job-id)")
        if job.get("file_id"):
            return job["file_id"]
        existing = {j["file_id"] for j in data["jobs"] if j.get("file_id")}
        max_n = 10 ** FILE_ID_DIGITS
        for _ in range(200):
            candidate = f"{random.randint(0, max_n - 1):0{FILE_ID_DIGITS}d}"
            if candidate not in existing:
                job["file_id"] = candidate
                return candidate
        raise RuntimeError("could not find a free 5-digit file_id after 200 tries")


def _unique_dest(company: str, suffix_digits: int, ext: str, tag: str = "") -> Path:
    """Fallback for the rare case tracker.py add is called without
    --job-id (so there's no jobs.json record to persist a file_id onto) -
    a random numeric suffix, retried on collision, still keeps files
    distinct without changing the human-readable company-name-first
    layout. Every normal call path passes --job-id and uses
    get_or_create_file_id() instead, so both files of the same job share
    one number."""
    base = sanitize_filename(company)
    if tag:
        base = f"{base}_{tag}"
    max_n = 10 ** suffix_digits
    for _ in range(50):
        candidate = BY_COMPANY_DIR / f"{base}_{random.randint(0, max_n - 1):0{suffix_digits}d}{ext}"
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"could not find a free filename for {base} after 50 tries")


def _set_link_cell(ws, row: int, col: int, display: str, target: Path) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = display
    cell.hyperlink = f"file://{target.resolve()}"
    cell.style = "Hyperlink"


def cmd_add(args) -> None:
    # File copy / PDF render happen outside the tracker lock deliberately -
    # they don't touch the .xlsx at all, so there's no reason for another
    # tracker.py process to wait on the (potentially slower, PDF-rendering)
    # work here just to log its own row.
    file_id = get_or_create_file_id(args.job_id) if args.job_id else None

    resume_dest = None
    if args.resume_path and Path(args.resume_path).exists():
        BY_COMPANY_DIR.mkdir(parents=True, exist_ok=True)
        if file_id:
            resume_dest = BY_COMPANY_DIR / f"{sanitize_filename(args.company)}_resume_{file_id}.pdf"
        else:
            resume_dest = _unique_dest(args.company, 2, ".pdf", tag="resume")
        shutil.copyfile(args.resume_path, resume_dest)

    jd_path = None
    if args.jd_path and Path(args.jd_path).exists():
        BY_COMPANY_DIR.mkdir(parents=True, exist_ok=True)
        if file_id:
            jd_path = BY_COMPANY_DIR / f"{sanitize_filename(args.company)}_{file_id}.pdf"
        else:
            jd_path = _unique_dest(args.company, 6, ".pdf")
        text_to_pdf.convert(
            Path(args.jd_path).read_text(),
            jd_path,
            title=f"{args.company} - {args.role}",
        )

    with locked_tracker_for_write():
        wb = ensure_workbook()
        ws = wb[SHEET_NAME]
        ws.append([
            args.company,
            args.role,
            args.status,
            args.date_posted or "",
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            args.location or "",
            args.address or "",
            args.work_type or "",
            args.salary or "",
            args.source or "",
            args.url or "",
            "",
            "",
            args.notes or "",
        ])
        row_idx = ws.max_row
        if resume_dest:
            _set_link_cell(ws, row_idx, COLUMNS.index("Resume") + 1, resume_dest.name, resume_dest)
        if jd_path:
            _set_link_cell(ws, row_idx, COLUMNS.index("Job Description") + 1, jd_path.name, jd_path)
        wb.save(TRACKER_FILE)

    extra = ""
    if resume_dest:
        extra += f" (resume -> {resume_dest})"
    if jd_path:
        extra += f" (jd -> {jd_path})"
    print(f"added {args.company} | {args.role} | {args.status} -> {TRACKER_FILE}{extra}")


def cmd_update_status(args) -> None:
    """Backs the dashboard's "Mark Submitted" button - the agent never
    clicks Submit itself, so it can't know when a real submission
    happened. This lets the user's own manual action update the same row
    tracker.py add already created, matched by company (+ role if given,
    for the rare case a company somehow has more than one row)."""
    with locked_tracker_for_write():
        wb = ensure_workbook()
        ws = wb[SHEET_NAME]
        norm_target = normalize_company(args.company)
        company_col = COLUMNS.index("Company") + 1
        role_col = COLUMNS.index("Role") + 1
        status_col = COLUMNS.index("Status") + 1
        updated = 0
        for row in ws.iter_rows(min_row=2):
            company_val = row[company_col - 1].value
            if not company_val or normalize_company(company_val) != norm_target:
                continue
            if args.role and (row[role_col - 1].value or "") != args.role:
                continue
            row[status_col - 1].value = args.status
            updated += 1
        if updated == 0:
            print(f"no matching row found for company={args.company!r} role={args.role!r}")
            sys.exit(1)
        wb.save(TRACKER_FILE)
    print(f"updated {updated} row(s) for {args.company!r} -> status={args.status!r}")


def main() -> None:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_list = sub.add_parser("list-companies")
    p_list.add_argument("--out", default=None)
    p_list.set_defaults(func=cmd_list_companies)

    p_check = sub.add_parser("check")
    p_check.add_argument("--company", required=True)
    p_check.set_defaults(func=cmd_check)

    p_add = sub.add_parser("add")
    p_add.add_argument("--job-id", default=None,
                        help="jobs.json id - lets the resume and JD PDFs share one persistent "
                             "5-digit file_id (stored on the job record) instead of two unrelated "
                             "random numbers. Always pass this for a real job.")
    p_add.add_argument("--company", required=True)
    p_add.add_argument("--role", required=True)
    p_add.add_argument("--status", required=True)
    p_add.add_argument("--location", default=None)
    p_add.add_argument("--address", default=None)
    p_add.add_argument("--source", default=None)
    p_add.add_argument("--url", default=None)
    p_add.add_argument("--resume-path", default=None)
    p_add.add_argument("--jd-path", default=None)
    p_add.add_argument("--date-posted", default=None)
    p_add.add_argument("--work-type", default=None)
    p_add.add_argument("--salary", default=None)
    p_add.add_argument("--notes", default=None)
    p_add.set_defaults(func=cmd_add)

    p_update = sub.add_parser("update-status")
    p_update.add_argument("--company", required=True)
    p_update.add_argument("--role", default=None,
                           help="Optional exact-match filter, in case a company somehow has more than one row")
    p_update.add_argument("--status", required=True)
    p_update.set_defaults(func=cmd_update_status)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
